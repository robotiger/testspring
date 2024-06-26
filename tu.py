import os
import sys
import threading
import time
import datetime
import serial 
import logging
import numpy as np
import openpyxl as xl
import OPi.GPIO as g
from pymodbus.client import ModbusSerialClient
import sqlitedict
import ifcfg
import requests
from scipy.optimize import minimize
import logging


logging.basicConfig(
    format='%(threadName)s %(name)s %(levelname)s: %(message)s',
    level=logging.INFO)


Ycontact=27.5
Fkr=200 

status={"progress":0,"cycles_done":0,"to_do":"nothing","clength":0,"ckx":0,"shrink":0,"status":""}
config={}

grb=None
gpp=None
cmb=None
mrk=None

used_gpio_pins={"son":7,"ena":12,"idx":26,"yp":15,"ym":22}

config=sqlitedict.SqliteDict('config_tp.db')

#Переключение вывода на выход и выдача на него единицы
def on(pin):
    if type(pin)==type("str"):
        pin=used_gpio_pins.get(pin)
    if pin:
        g.setup(pin,g.OUT)
        g.output(pin,1)
        g.cleanup(pin)  
    
def off(pin):
    if type(pin)==type("str"):
        pin=used_gpio_pins.get(pin)
    if pin:    
        g.setup(pin,g.OUT)
        g.output(pin,0)
        g.cleanup(pin)  


def read_value(pin):
    val=None
    if type(pin)==type("str"):
        pin=used_gpio_pins.get(pin)
    if pin:    
        g.setup(pin,g.IN)
        val=g.input(pin)
        g.cleanup(pin)
    return val

def logInf(s): 
    logging.info(s)
    #print(s)
    


def setmb(adr,val):
    if cmb:
        cmb.write_register(adr,val,1)


def runmb(speed):
    runspeed=speed
    setmb(75,maxspeed)
    setmb(76,min(speed//2,maxspeed))
    setmb(0x1010,0x1234)
    setmb(0x1010,0x2222)
    time.sleep(0.5)
    setmb(76,min(speed,maxspeed))

# A class that extends the Thread class
class gp(threading.Thread):
    def __init__(self,stop_event,pin):
        threading.Thread.__init__(self)
        self.stop_event=stop_event
        
        if type(pin)==type("str"):
            self.name=pin
            pin=used_gpio_pins.get(pin)        
        self.pin=pin
        g.setwarnings(False)
        g.setmode(g.BOARD)
        g.setup(self.pin,g.IN)
        self.count=0
        self.tlp=time.time()
        self.last=0
        self.stop=0
        self.freq=0
        self.callback_stop=None
    def run(self):
        while(not self.stop_event.is_set()):
            self.last<<=1
            self.last|=g.input(self.pin)
            self.last&=0xff
            if self.last==0x7:
                if self.name=='idx':
                    self.count-=1
                else:
                    self.count=0
                if self.count<=0:
                    if self.callback_stop :
                        self.callback_stop()                    
            time.sleep(0.001)
    def read_value(self):
            return g.input(self.pin)

class grbs(threading.Thread):
    def __init__(self,stop_event,devs):
        logInf("grbs __init__ ")
        threading.Thread.__init__(self)  
        self.stop_event=stop_event
        self.port=devs['grbl']
        self.pack=b''
        self.length=1
        self.grblserial=None
    def run(self): 
        logInf("run grbs")
        with serial.Serial(self.port, 115200, timeout=1) as self.grblserial:
            buf=b''
            while(not self.stop_event.is_set()):
                logInf('grbl port is',self.grblserial._port)
                tmp=self.grblserial.readline()
                if len(tmp)>0:
                    self.buf=tmp.decode()
                    logInf(f'grbs rcv {buf}')
        self.grblserial=None
                    
    def write(self,data):
        if type(data)==type('str'):
            data=data.encode()
        if self.grblserial:
            x=self.grblserial.write(data)
            logInf(f"sent to grbl {x} byte: {data}")
        
    def soft_reset(self):
        self.write(b'!')
        setmb(76,100)
        time.sleep(1)
        self.write(b'\x18')
        setmb(0x1010,0)
        logInf("soft reset done")


class mark(threading.Thread):

    def __init__(self,stop_event,port):
        logInf("mark __init__")
        threading.Thread.__init__(self)  
        self.stop_event=stop_event
        self.ok=False        
        self.buf=''
        self.port=port['mark']

    def run(self): 
        logInf("run mark")
        
        val=None
        with serial.Serial(self.port, 115200, timeout=1) as self.ser:
            while(not self.stop_event.is_set()):
                self.buf=self.ser.readline().decode()
                logInf("mark read",self.buf)            
            
        logInf("Mark closed!!!")
        

    def write(self,data):
        self.ser.write(data)
    
    def ask(self):
        firstMeasure=0
        for i in range(6):
            time.sleep(0.5)
            self.ser.write(b'?\r')
            time.sleep(0.5)
            measure=float(self.buf)
            if measure > Fkr:
                grb.soft_reset()
                
            if abs(firstMeasure-measure)<0.1:
                break
            firstMeasure=measure
        return measure

class measures(threading.Thread):
    
    def __init__(self,stop_event):
        threading.Thread.__init__(self)
        self.stop_event=stop_event   
        self.atHome=False
        self.sx=[]
        
    def find_edge(self):
        logInf("start find edge")
        on("son")
        for t in range(5):
            
            gpIdx.count=1
            gpIdx.stop=0
            grb.write(b"g91g1f1000x1000\n")
            time.sleep(5)
            if gpIdx.count==0:
                break
                #print(gpIdx.stop,gpIdx.count,gpIdx.last)
        # если не вал не повернулся вернём ошибку        
        return gpIdx.count==1
        #logInf("edge found rotation stop at idx")
        #time.sleep(2)
    
    def runtest(self,speed,count):
        global status
        gpIdx.stop=0
        gpIdx.count=count
        runmb(speed)
        while(gpIdx.count>0):
            #print(gpIdx.stop,gpIdx.count,gpIdx.last)
            if status['to_do']=='stoptest':
                break               
            time.sleep(0.5)        
        runmb(0)

    def runmesure(self):
        self.forces=[]
        logInf("move y ")
        off("ena")
        
        grb.write(f"g91g1f1000y{Ycontact+config['lmin']-config['lstep']}\n")
        #input('pause before mesure')
        time.sleep(0.1)
        force=mrk.ask()
        for i in self.sx:
            grb.write(f"g91g1f1000y{config['lstep']}\n".encode())
            time.sleep(0.1)
            force=mrk.ask()
            self.forces.append(force) 

            logInf(f" dist {i}, force {force}")

        grb.write(f"g91g1f1000y-{config['lmax']}\n".encode())
        time.sleep(5)
        #on("ena")
        return self.forces

    
    
    
    def home_ym(self):
        self.atHome=False
        off("ena")
        time.sleep(5)
        if not gpYm.read_value(): #если не на датчике наедем на него
            grb.write("g91g21g1f1000y-60\n") #
            time.sleep(10)
        #останавливается самостоятельно по soft_reset
    
        for i in range(20):
            if gpYm.read_value(): #уже за датчиком, нужно сойти с датчика
                grb.write("g91g21g1f1000y1\n") #сходим на 1 мм
                time.sleep(1)
                if mrk.ask()>Fkr:
                    grb.write("g91g21g1f1000y-1\n")
                    
            else:
                break #как только сошли прекращаем движение
                
        grb.write("g91g21g1f1000y0.3\n") 
        time.sleep(0.5)    
        if gpYm.read_value(): #не сошли с датчика. ошибка
            logInf("не сошли с датчика. ошибка")
            #stop_event.set()
        else:    
            grb.write("g91g21g1f10y-3\n")
            time.sleep(10)
            if gpYm.read_value():
                logInf("по оси Y вышли в ноль по датчику (ym)")
                self.atHome=True
            else:
                logInf("датчик ym не нашли")
                #stop_event.set()
        return self.atHome
                                
    def run_test(self):
        off("ena")
        on("son")
    
        time.sleep(5)
    
        self.xlMakeHeader()
        self.home_ym()
        if not self.atHome:
            self.home_ym()
        self.cycles=0
        #input('pause')
        while(True):
            runspeed= int(config["freq"]*38*60/17)

            if status['to_do']=='stoptest':
                break             
            runcycles=min(config["cyclesbetween"],config["cycles"]-config['cycles_complete'])
            self.runtest(runspeed,runcycles)
            time.sleep(2)
            if status['to_do']=='stoptest':
                break                 
            self.find_edge()
            
            self.xlSaveRow(self.runmesure())
            
            if status['to_do']=='stoptest':
                break            
            self.home_ym()
            if status['to_do']=='stoptest':
                break                 
        
            if config['cycles_complete']>=config["cycles"]:
                status['to_do']='nothing'
                break
            if stop_event.is_set():
                break
        time.sleep(1)
        
        on("ena")
        off("son")        
        
        
    def xlMakeHeader(self):
        if os.path.exists(config['xlfilename']):
            wb=xl.load_workbook(config['xlfilename'])
        else:
            wb=xl.Workbook()
        
        ws=wb.active
        
        ws.cell(row=1,column=2,value='Протокол тестирования пружины')
        ws.cell(row=1,column=1,value=datetime.datetime.now())
        
        row=2

        
        for t in tab:
            ws.cell(row=row,column=1,value=tab[t])
            ws.cell(row=row,column=5,value=config[t])
            row+=1
        
        ws.cell(row=row,column=1,value='Дата')
        ws.cell(row=row,column=2,value='Циклы')
        ws.cell(row=row,column=3,value='Длина')
        ws.cell(row=row,column=4,value='Кх')
        ws.cell(row=row,column=5,value='Усадка')
    
        column=6
        self.sx=list(range(config['lmin'],config['lmax']+config['lstep'],config['lstep']))
        
        for i in self.sx: 
            #числа только целые пока
            ws.cell(row=row,column=column,value=i)
            column+=1  
        config['startrow']=row+1    
        wb.save(config['xlfilename'])
    
    def xlSaveRow(self,forces):
        global status
        wb=xl.load_workbook(config['xlfilename'])
        ws=wb.active
        
        config['cycles_complete']+=config["cyclesbetween"]
        gpIdx.count=config["cyclesbetween"] #чтобы прогресс не скакал а двигался плавно
        mc=config['startrow']
        
        self.sx=list(range(config['lmin'],config['lmax']+config['lstep'],config['lstep']))
        sx=np.array(self.sx)
        sf=np.array(self.forces)
        
        #немного расчетов
        def f(x,sx,sf):
            return ((sx*x[0]+x[1]-sf)**2).sum()
        res=minimize(f,x0=[3,3],args=(sx,sf))
        ckx=res.x[0]
        clength=res.x[1]/res.x[0]+config["ldistance"]
        # сохраняем результаты расчета в статус для обновления веб страницы
        status['ckx']=ckx
        status['clength']=clength
        status['shrink']=config['slength']-clength

        #заполняем строку данными
        ws.cell(row=mc,column=1,value=datetime.datetime.now()) 
        ws.cell(row=mc,column=2,value=config['cycles_complete']) 
        ws.cell(row=mc,column=3,value=clength) 
        ws.cell(row=mc,column=4,value=ckx) 
        ws.cell(row=mc,column=5,value=config['slength']-clength) 
        
        
        for i in range(len(forces)):
            ws.cell(row=mc,column=i+6,value=self.forces[i]) 
            
        config['startrow']+=1
        wb.save(config['xlfilename'])

class webrun(threading.Thread):
    def __init__(self,stop_event):
        threading.Thread.__init__(self)
        self.stop_event=stop_event

    def run(self):
        global status
        global config
        res_ok=False
        while(not self.stop_event.is_set()):
            try:
                res=requests.get('http://localhost:5000/status')
                if res.ok:
                    newstatus=res.json()
                    res_ok=res.ok
                status['to_do']=newstatus['to_do']
                for x in tab:
                    config[x]=newstatus[x]
                #print(f"got status to_do {newstatus['to_do']} \n{config}")
                #print('.',end='')
            except:
                #print('-',end='')
                #print("отключено приложение веб")
                pass
            if res_ok:
                
                if gpIdx.count>=0:
                    status['cycles_done']=config["cycles_complete"]+config["cyclesbetween"]-gpIdx.count
                else:
                    status['cycles_done']=0
                    
                if status['cycles_done']>0:
                    status['progress']=status['cycles_done']*100//config["cycles"]
                else:
                    status['progress']=0
                try:
                    rr=requests.post('http://localhost:5000/status',json=status) 
                    #print(f"sent status {status}")
                    
                except:
                    logInf("отключено приложение веб")            
            time.sleep(1)
    
def scanUSB():
    devs={}
    for portn in range(3):
        portname=f'/dev/ttyUSB{portn}'
        logInf(f'try open {portname}')
        try:
            with serial.Serial(portname, 115200, timeout=1) as ser:    
                
                time.sleep(1)
                ser.readline()
                s=ser.readline().decode()
                print(s)
                if len(s)>0:
                    if s[:4]=="Grbl":
                        devs["grbl"]=portname
                ser.write(b'?\r')
                time.sleep(1)
                s=ser.readline().decode()
                if s[:1]=='0':
                    devs["mark"]=portname
        except:
            pass
    return devs


# ****************main ********************
if __name__ == '__main__':


    stop_event = threading.Event()
    
    devs=scanUSB()
    
    if len(devs)<2:
        print(f'devs {devs} не достаточно')
        exit()
    logInf(devs)
    time.sleep(10)
    
   
    gpIdx=gp(stop_event,"idx")
    gpIdx.start()
    
    gpYp=gp(stop_event,"yp")
    gpYp.start()
    
    gpYm=gp(stop_event,"ym")
    gpYm.start()
    
    grb=grbs(stop_event,devs)
    grb.start()    
    
    
    gpIdx.callback_stop=grb.soft_reset
    gpYm.callback_stop=grb.soft_reset
    gpYp.callback_stop=grb.soft_reset    
    
    cmb=ModbusSerialClient('/dev/ttyS1',parity='E')
    cmb.connect()
    cmb.write_register(0x1010,0x0,1)    
    

    mrk=mark(stop_event,devs)
    mrk.start()
    
    ms=measures(stop_event)
    
    web=webrun(stop_event)
    web.start()    
    
    for i in range(1000):
        ms.home_ym()
        logInf(mrk.ask())
        ms.runmesure()
        grb.write("g91g21g1f1000y1\n") 
        time.sleep(1)
        grb.write("g91g21g1f1000y-1\n") 
        time.sleep(1)
        