import os
import sys
import threading
import time
import datetime
import serial 
import logging
import numpy as np
import openpyxl as xl
import OPi.GPIO as g
from pymodbus.client import ModbusSerialClient
import sqlitedict
import ifcfg
import requests
from scipy.optimize import minimize



logging.basicConfig(
    format='%(threadName)s %(name)s %(levelname)s: %(message)s',
    level=logging.INFO)



config=sqlitedict.SqliteDict('config_tp.db')
config.autocommit=True

logname='/home/bfg/data/ts.log'

tab={
 'snum': 'Номер теста',
 'sname': 'Наименование',
 'slength': 'Длина',
 'sdiameter': 'Диаметр пружины',
 'sdp': 'Толщина прутка',
 'snrot': 'Число витков',
 'smatherial': 'Материал',
 'skxnom': 'Номинальный коэфф жесткости',
 'cycles': 'Число циклов сжатия',
 'cyclesbetween': 'Число циклов сжатия между измерениями',
 'freq': 'Частота сжатия',
 'lmin': 'Начальное сжатие',
 'lmax': 'Максимальное сжатие',
 'lstep': 'Шаг измерения усилия пружины',
 'ldistance':'Длина поджатой пружины',
 'xlfilename':'Имя файла протокола',
 #'cycles_complete':'Выполнено циклов'
 }
    
#Yfirststep=5
#Ystep=1
#Ymax=24

Ycontact=27.5
Fkr=200 

#totalCycles=200
#mesureCycles=20
maxspeed=2500


status={"progress":0,"cycles_done":0,"to_do":"nothing","clength":0,"ckx":0,"shrink":0,"status":""}
config={}

grb=None
gpp=None
cmb=None
mrk=None

used_gpio_pins={"son":7,"ena":12,"idx":26,"yp":15,"ym":22}


def lprint(s): 
    logging.info(s)
    #loglenmax=10000000   
    #logging.basicConfig(format='%(asctime)s | %(message)s', datefmt='%Y/%m/%d %H: %M: %S ', filename=logname, level=logging.INFO)
    #logging.info(s)
    
    #app.logger.info(s)
    #try: 
        #statinfo= os.stat(logname)

        #if statinfo.st_size >loglenmax: 
            #os.rename(logname, lognameold)
            #logging.shutdown()
            #logging.basicConfig(format='%(asctime)s | %(message)s', datefmt='%Y/%m/%d %H: %M: %S ', filename=logname, level=logging.INFO)
    #except: 
        #pass



#Переключение вывода на выход и выдача на него единицы
def on(pin):
    if type(pin)==type("str"):
        pin=used_gpio_pins.get(pin)
    if pin:
        g.setup(pin,g.OUT)
        g.output(pin,1)
        g.cleanup(pin)  
    
def off(pin):
    if type(pin)==type("str"):
        pin=used_gpio_pins.get(pin)
    if pin:    
        g.setup(pin,g.OUT)
        g.output(pin,0)
        g.cleanup(pin)  


def read_value(pin):
    val=None
    if type(pin)==type("str"):
        pin=used_gpio_pins.get(pin)
    if pin:    
        g.setup(pin,g.IN)
        val=g.input(pin)
        g.cleanup(pin)
    return val


def setmb(adr,val):
    if cmb:
        cmb.write_register(adr,val,1)


def runmb(speed):
    runspeed=speed
    setmb(75,maxspeed)
    setmb(76,min(speed//2,maxspeed))
    setmb(0x1010,0x1234)
    setmb(0x1010,0x2222)
    time.sleep(0.5)
    setmb(76,min(speed,maxspeed))


# A class that extends the Thread class
class gp(threading.Thread):
    def __init__(self,stop_event,pin):
        threading.Thread.__init__(self)
        self.stop_event=stop_event
        
        if type(pin)==type("str"):
            self.name=pin
            pin=used_gpio_pins.get(pin)        
        self.pin=pin
        g.setwarnings(False)
        g.setmode(g.BOARD)
        g.setup(self.pin,g.IN)
        self.count=0
        self.tlp=time.time()
        self.last=0
        self.stop=0
        self.freq=0
        self.callback_stop=None
    def run(self):
        while(not self.stop_event.is_set()):
            self.last<<=1
            self.last|=g.input(self.pin)
            self.last&=0xff
            if self.last==0x7:
                if self.name=='idx':
                    self.count-=1
                else:
                    self.count=0
                if self.count<=0:
                    if self.callback_stop :
                        self.callback_stop()                    
            time.sleep(0.001)
    def read_value(self):
        return g.input(self.pin)
        
class mark(threading.Thread):
    def __init__(self,stop_event,port):
        lprint("mark __init__")
        threading.Thread.__init__(self)  
        self.stop_event=stop_event
        self.ok=False        
        self.buf=''
        self.port=port['mark']
    def run(self): 
        lprint("run mark")
        with serial.Serial(self.port, 115200, timeout=1) as self.markserial:
            while(not self.stop_event.is_set()):
                #lprint(f'mark port is {self.markserial._port}')
                tmp=self.markserial.readline()
                if len(tmp)>0:
                    self.buf=tmp.decode().strip()
                    if self.buf=="OVERRANGE":
                        self.buf="500.5"
                    #lprint(f'mark read {self.buf}')            
        lprint("Mark closed!!!")
    def ask(self):
        firstMeasure=0
        measure=-1
        for i in range(6):
            time.sleep(0.5)
            self.markserial.write(b'?\r')
            time.sleep(0.5)
            if len(self.buf)>0:
                measure=float(self.buf)
            if measure > Fkr:
                grb.soft_reset()
            if abs(firstMeasure-measure)<0.1:
                break
            firstMeasure=measure
        return measure

class grbs(threading.Thread):
    def __init__(self,stop_event,devs):
        lprint("grbs __init__ ")
        threading.Thread.__init__(self)  
        self.stop_event=stop_event
        self.port=devs['grbl']
        self.buf=''
        self.length=1
    def run(self): 
        lprint("run grbs")
        with serial.Serial(self.port, 115200, timeout=1) as self.grblserial:
            buf=b''
            while(not self.stop_event.is_set()):
                #print(f'grbl port is {self.grblserial._port}')
                tmp=self.grblserial.readline()
                if len(tmp)>0:
                    self.buf=tmp.decode()
                    #lprint(f'grbs rcv {buf}')
                    
    def write(self,data):
        if type(data)==type('str'):
            data=data.encode()
        x=self.grblserial.write(data) # sw
        lprint(f"sent to grbl {x} byte: {data}")
        
    def soft_reset(self):
        lprint("soft reset ")
        self.write(b'!') #sw
        setmb(76,100)
        time.sleep(1)
        self.write(b'\x18')#sw
        setmb(0x1010,0)
        lprint("soft reset done")





class measures(): 
    

        
    def find_edge(self):# Поворот эксцентрика в исходное
        lprint("start find edge Поворот эксцентрика в исходное")
        on("son")
        for t in range(5):
            
            gpIdx.count=1
            gpIdx.stop=0
            grb.write(b"g91g1f1000x1000\n")
            time.sleep(5)
            if gpIdx.count==0:
                break
                #print(gpIdx.stop,gpIdx.count,gpIdx.last)
        # если не вал не повернулся вернём ошибку  
        time.sleep(5)
        off("son")
        return gpIdx.count==0 
        #lprint("edge found rotation stop at idx")
        #time.sleep(2)
        
    def home_ym(self):
        self.atHome=False
        lprint("Выход в ноль Y координаты")
        off("ena")
        on("son")
        time.sleep(5)
        if not gpYm.read_value(): #если не на датчике наедем на него
            lprint("двигаемся на датчик Y минус")
            grb.write("g91g21g1f1000y-60\n") #
            time.sleep(10)
        #останавливается самостоятельно по soft_reset
    
        for i in range(20):
            if gpYm.read_value(): #уже за датчиком, нужно сойти с датчика
                print("двигаемся с датчика Y минус")
                grb.write("g91g21g1f1000y1\n") #сходим на 1 мм
                time.sleep(1)
                #if mrk.ask()>Fkr:
                #    grb.write("g91g21g1f1000y-1\n")
                # надо вернуть   
            else:
                print("съехали с датчика Y минус")
                break #как только сошли прекращаем движение
                
        grb.write("g91g21g1f1000y0.3\n") 
        time.sleep(0.5)    
        if gpYm.read_value(): #не сошли с датчика. ошибка
            lprint("не сошли с датчика. ошибка")
            #stop_event.set()
        else:    
            grb.write("g91g21g1f10y-3\n")
            time.sleep(10)
            if gpYm.read_value():
                lprint("по оси Y вышли в ноль по датчику (ym)")
                self.atHome=True
            else:
                lprint("датчик ym не нашли")
                self.atHome=False
                #stop_event.set()
        return self.atHome        
    
    def runtest(self,speed,count):
        global status
        on("son")
        off("ena")
        #Перед вращением  отведем измеритель в крайнее левое положение
        if not gpYm.read_value(): #если не на датчике наедем на него
            lprint("двигаемся в сторону датчика Y минус на 60мм")
            grb.write("g91g21g1f1000y-60\n") 
            
        gpIdx.stop=0
        gpIdx.count=count
        runmb(speed)
        while(gpIdx.count>0):
            #print(gpIdx.stop,gpIdx.count,gpIdx.last)
            if status['to_do']=='stoptest':
                break               
            time.sleep(0.1)        
        runmb(0)
        

    def runmesure(self):
        global status
        off("ena")  #включим привод Y      
        #повернём эксцентрик в исходное
        x_home=self.find_edge()
        
        #Перед измерением Выйдем в ноль Y 
        y_home=self.home_ym()
        
        if x_home and y_home:
            self.forces=[]

            off("son") #Перед измерением отключим привод X
            lprint("move y to first mesuring distance")       
            grb.write(f"g91g1f1000y{Ycontact+config['lmin']-config['lstep']}\n")
            #Встаем в положение на один шаг меньше чем первое измерение
            #в цикле двигаем на шаг затем измеряем
            time.sleep(0.5)
            force=mrk.ask()
            self.sx=list(range(config['lmin'],config['lmax']+config['lstep'],config['lstep']))            
            for i in self.sx:
                
                grb.write(f"g91g1f1000y{config['lstep']}\n".encode())
                time.sleep(0.5)
                force=mrk.ask()
                self.forces.append(force) 
    
                lprint(f" dist {i}, force {force}")
            
            #Отводим Y в исходное положение
            grb.write(f"g91g1f1000y-{Ycontact+config['lmax']}\n".encode())
            time.sleep(5)
            on("son") #после измерения включим привод X
    
            status["forces"]=self.forces
            return True
        else:            
            lprint(f"Не вышли в ноль до начала измерения x {x_home} y {y_home}")
            return False

    
                                
    def run_test(self): #Основной цикл тестирования измерения
        off("ena")
        on("son")
    
        time.sleep(5)
        if status['to_do']=='stoptest':
            return()
        
        self.xlMakeHeader()
        config['cycles_complete']=0
        
        self.runmesure()
        self.xlSaveRow()        


        while(config['cycles_complete']<=config["cycles"]):
            lprint(f"{config['cycles_complete']}<{config['cycles']}")      
            
            runspeed= int(config["freq"]*38*60/17)

            if status['to_do']=='stoptest':
                break             

            runcycles=min(config["cyclesbetween"],config["cycles"]-config['cycles_complete'])
            self.runtest(runspeed,runcycles) #
            time.sleep(2)
            
            if status['to_do']=='stoptest':
                break                 
            
            config['cycles_complete']+=config["cyclesbetween"]
            gpIdx.count=config["cyclesbetween"] #чтобы прогресс не скакал а двигался плавно            
            
            self.runmesure() # выполнить измерения
            self.xlSaveRow() # сохранить строку в эксельку
            
            if status['to_do']=='stoptest':
                break            

            if stop_event.is_set():
                break
        time.sleep(1)
        
        on("ena")
        off("son")        
        
        
    def xlMakeHeader(self):
        if os.path.exists(config['xlfilename']):
            wb=xl.load_workbook(config['xlfilename'])
        else:
            wb=xl.Workbook()
        
        ws=wb.active
        
        ws.cell(row=1,column=2,value='Протокол тестирования пружины')
        ws.cell(row=1,column=1,value=datetime.datetime.now())
        
        row=2

        for t in tab:
            ws.cell(row=row,column=1,value=tab[t])
            ws.cell(row=row,column=5,value=config[t])
            row+=1
        
        ws.cell(row=row,column=1,value='Дата')
        ws.cell(row=row,column=2,value='Циклы')
        ws.cell(row=row,column=3,value='Длина')
        ws.cell(row=row,column=4,value='Кх')
        ws.cell(row=row,column=5,value='Усадка')
    
        column=6
        self.sx=list(range(config['lmin'],config['lmax']+config['lstep'],config['lstep']))
        
        for i in self.sx: 
            #числа только целые пока
            ws.cell(row=row,column=column,value=i)
            column+=1  
        config['startrow']=row+1    
        wb.save(config['xlfilename'])
    
    def xlSaveRow(self):
        global status
        wb=xl.load_workbook(config['xlfilename'])
        ws=wb.active
        

        mc=config['startrow']
        
        self.sx=list(range(config['lmin'],config['lmax']+config['lstep'],config['lstep']))
        sx=np.array(self.sx)
        sf=np.array(self.forces)
        
        #немного расчетов
        def f(x,sx,sf):
            return ((sx*x[0]+x[1]-sf)**2).sum() 
        res=minimize(f,x0=[3,3],args=(sx,sf))
        ckx=res.x[0]
        clength=res.x[1]/res.x[0]+config["ldistance"]
        # сохраняем результаты расчета в статус для обновления веб страницы
        status['ckx']=ckx
        status['clength']=clength
        status['shrink']=config['slength']-clength

        #заполняем строку данными
        ws.cell(row=mc,column=1,value=datetime.datetime.now()) 
        ws.cell(row=mc,column=2,value=config['cycles_complete']) 
        ws.cell(row=mc,column=3,value=clength) 
        ws.cell(row=mc,column=4,value=ckx) 
        ws.cell(row=mc,column=5,value=config['slength']-clength) 
        
        
        for i in range(len(self.forces)):
            ws.cell(row=mc,column=i+6,value=self.forces[i]) 
            
        config['startrow']+=1
        wb.save(config['xlfilename'])
        


class webrun(threading.Thread):
    def __init__(self,stop_event):
        threading.Thread.__init__(self)
        self.stop_event=stop_event

    def run(self):
        global status
        global config
        res_ok=False
        while(not self.stop_event.is_set()):
            try:
                res=requests.get('http://localhost:5000/status')
                if res.ok:
                    newstatus=res.json()
                    res_ok=res.ok
                status['to_do']=newstatus['to_do']
                for x in tab:
                    config[x]=newstatus[x]
                #lprint(f"got status to_do {newstatus['to_do']} \n{config}")
                #print('.',end='')
            except:
                #print('-',end='')
                lprint("отключено приложение веб")
            if res_ok:
                
                if gpIdx.count>=0:
                    try:
                        status['cycles_done']=config["cycles_complete"]+config["cyclesbetween"]-gpIdx.count
                    except:
                        pass
                else:
                    status['cycles_done']=0
                    
                if status['cycles_done']>0:
                    status['progress']=status['cycles_done']*100//config["cycles"]
                else:
                    status['progress']=0
                try:
                    rr=requests.post('http://localhost:5000/status',json=status) 
                    #print(f"sent status {status}")
                    
                except:
                    lprint("отключено приложение веб")            
            time.sleep(1)
    

def scanUSB():
    devs={}
    for portn in range(3):
        portname=f'/dev/ttyUSB{portn}'
        lprint(f'try open {portname}')
        try:
            with serial.Serial(portname, 115200, timeout=1) as ser:    
                
                time.sleep(1)
                ser.readline()
                s=ser.readline().decode()
                lprint(s)
                if len(s)>0:
                    if s[:4]=="Grbl":
                        devs["grbl"]=portname
                ser.write(b'?\r')
                time.sleep(1)
                s=ser.readline().decode()
                if s[:1]=='0':
                    devs["mark"]=portname
        except:
            pass
    return devs
                
                
                


# ****************main ********************
if __name__ == '__main__':

    on("ena")
    off("son")    

    stop_event = threading.Event()
    
    devs=scanUSB()
    
    if not 'mark' in devs:
        lprint(f'mark-10 не подключен')
        exit()
    lprint(repr(devs))

    mrk=mark(stop_event,devs)
    mrk.start()
    
    grb=grbs(stop_event,{'grbl':'/dev/ttyS2'})
    grb.start()
    
    cmb=ModbusSerialClient('/dev/ttyS1',parity='E')
    cmb.connect()
    cmb.write_register(0x1010,0x0,1)
    
    
    gpIdx=gp(stop_event,"idx")
    gpIdx.start()
    
    gpYp=gp(stop_event,"yp")
    gpYp.start()
    
    gpYm=gp(stop_event,"ym")
    gpYm.start()
    
    gpIdx.callback_stop=grb.soft_reset
    gpYm.callback_stop=grb.soft_reset
    gpYp.callback_stop=grb.soft_reset



    time.sleep(10)
    #lprint(f'mark run is {mrk.ok}')
    

    ms=measures()

    web=webrun(stop_event)
    web.start()
    

    while(True):

        #lprint(f"main {status['to_do']} \n status {status}\n config {config}")
        
        if status['to_do']=='setspring':
            lprint(" find edge")
            ms.find_edge()
            
            lprint("go home ")
            if ms.home_ym():
                status['YatHome']='athome'
            else:
                status['YatHome']='notathome'
                    
        if status['to_do']=='runtest':
            
            #if config['cycles_complete']<config['cycles']:
                ms.run_test()

            

        if status['to_do']=='rtest':
            ms.find_edge()
 
        if status['to_do']=='htest':
            ms.home_ym() 
            
        if status['to_do']=='ktest':
            runspeed= int(config["freq"]*38*60/17)
            ms.runtest(runspeed,config['cyclesbetween'])
   
        if status['to_do']=='mtest':
            ms.runmesure()

        time.sleep(1)
        


def joins():
    grb.join()
    gpIdx.join()
    gpYm.join()
    gpYp.join()
    #mrk.join()



